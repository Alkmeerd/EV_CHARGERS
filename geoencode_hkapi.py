import openpyxl, requests
from urllib.parse import quote
import time

filepath = "/Users/lamderek/Documents/Sudo/Study/University Study/Year 4/RECO4001 Dissertation/01_EV Charging Stations/11_Analysis"

wb = openpyxl.load_workbook(filepath + "/Charger_Locations_v3.xlsx")
sht = wb["2021-09"]

print("workbook loaded")


# for j in range(5,564):
#     val = sht.cell(row = j, column = 3).value
#     for i in range(len(val)):
#         if val[i] > u'\u4e00' and val[i] < u'\u9fff':
#             sht.cell(row = j, column = 4).value = val[0:i]
#             sht.cell(row = j, column = 5).value = val[i:len(val)]
#             break

# wb.save(filepath + "/Charger_Locations_v2.xlsx")

discrepancies = []
error_name = []

for i in range(5,564):
    try:
        eng_name = sht.cell(row = i, column = 4).value

        api_endpoint = 'https://geodata.gov.hk/gs/api/v1.0.0/locationSearch?q='
        querystring = quote(eng_name)
        response = requests.get(api_endpoint + querystring)

        # success retrieval
        if response.status_code == 200:
            # assumption: only 1 match
            address_eng = response.json()[0]['addressEN']
            northing_eng = response.json()[0]['x']
            easting_eng = response.json()[0]['y']

        else:
            print(response.status_code)
            error_name.append(i)

        time.sleep(0.01)

        if (i % 20 == 0):
            print(i, "completed")

        sht.cell(row = i, column = 6).value = northing_eng
        sht.cell(row = i, column = 7).value = easting_eng
    
    except:
        print(i, "ERRORERRORERROR", "Status Code = ", response.status_code)
        error_name.append(i)

    
print("ERROR:", error_name)
wb.save(filepath + "/Charger_Locations_v4.xlsx")


